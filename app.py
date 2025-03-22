from flask import Flask, render_template, request, redirect, url_for, flash, send_file
from flask_sqlalchemy import SQLAlchemy
from flask_login import LoginManager, UserMixin, login_user, login_required, logout_user, current_user
from functools import wraps
import os
from datetime import datetime
import csv
import io
import tempfile
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from models import db, User, Department, Course, Classroom, Schedule
from sqlalchemy import inspect, text

# =====================================================================================
# Ders Programı Yönetim Sistemi
# Bu sistem üniversite için bir ders programı yönetimi sağlar.
# Özellikler:
# - Bölüm ekleme, silme ve düzenleme
# - Ders ekleme, silme ve düzenleme
# - Derslik ekleme, silme ve düzenleme
# - Kullanıcı yönetimi (admin, öğretim görevlisi, öğrenci)
# - Ders programı oluşturma ve Excel'e aktarma
# =====================================================================================

# Göreceli yolları kullanarak dizinleri belirle
TEMPLATE_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'templates')
DB_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'ders_programi.db')

# Flask uygulamasını oluştur ve yapılandır
app = Flask(__name__, template_folder=TEMPLATE_DIR)
app.config['SECRET_KEY'] = 'gizli-anahtar-buraya'  # Güvenlik için session anahtarı
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///' + DB_PATH  # SQLite veritabanı bağlantısı
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False  # Performans için takip özelliğini kapat

# Veritabanı ve giriş yöneticisini başlat
db.init_app(app)
login_manager = LoginManager(app)
login_manager.login_view = 'login'  # Giriş yapılmadığında yönlendirilecek sayfa

# Flask-Login için kullanıcı yükleme fonksiyonu
@login_manager.user_loader
def load_user(user_id):
    """
    Flask-Login için kullanıcı kimliğinden kullanıcı nesnesini yükler
    :param user_id: Kullanıcı kimlik numarası
    :return: Kullanıcı nesnesi veya None
    """
    return User.query.get(int(user_id))

# Admin yetkisi gerektiren sayfalar için dekoratör
def admin_required(f):
    """
    Bir rotaya sadece admin kullanıcıların erişebilmesini sağlayan dekoratör
    :param f: Dekore edilecek fonksiyon
    :return: Dekore edilmiş fonksiyon
    """
    @wraps(f)
    def decorated_function(*args, **kwargs):
        # Kullanıcı giriş yapmamış veya admin değilse erişimi engelle
        if not current_user.is_authenticated or current_user.role != 'admin':
            flash('Bu sayfaya erişim yetkiniz yok!', 'error')
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

# Ana sayfa - Ders programına yönlendirir
@app.route('/')
def index():
    """
    Ana sayfa, kullanıcıyı ders programı görüntüleme sayfasına yönlendirir
    """
    return redirect(url_for('view_schedule'))

# Giriş sayfası
@app.route('/login', methods=['GET', 'POST'])
def login():
    """
    Kullanıcı giriş sayfası
    GET: Giriş formunu göster
    POST: Kullanıcı giriş bilgilerini kontrol et
    """
    if request.method == 'POST':
        # Form verilerini al
        username = request.form.get('username')
        password = request.form.get('password')
        # Kullanıcıyı veritabanında ara
        user = User.query.filter_by(username=username).first()
        
        # Kullanıcı varsa ve şifre doğruysa giriş yap
        if user and user.password == password:
            login_user(user)
            return redirect(url_for('view_schedule'))
        
        # Giriş başarısızsa hata mesajı göster
        flash('Geçersiz kullanıcı adı veya şifre!', 'error')
    return render_template('login.html')

# Çıkış sayfası
@app.route('/logout')
@login_required  # Sadece giriş yapmış kullanıcılar çıkış yapabilir
def logout():
    """
    Kullanıcının sistemden çıkış yapmasını sağlar
    """
    logout_user()
    return redirect(url_for('login'))

# Bölümler sayfası
@app.route('/departments', methods=['GET', 'POST'])
@admin_required  # Sadece adminler bölüm ekleyip silebilir
def departments():
    """
    Bölüm yönetim sayfası
    GET: Bölüm listesini göster
    POST: Yeni bölüm ekle
    """
    if request.method == 'POST':
        # Form verilerini al
        code = request.form.get('code')
        name = request.form.get('name')
        
        # Aynı kodla başka bölüm var mı kontrol et
        if Department.query.filter_by(code=code).first():
            flash('Bu bölüm kodu zaten kullanımda!', 'error')
            return redirect(url_for('departments'))
        
        # Yeni bölüm oluştur ve kaydet
        department = Department(code=code, name=name)
        db.session.add(department)
        db.session.commit()
        
        flash('Bölüm başarıyla eklendi!', 'success')
        return redirect(url_for('departments'))
    
    # Tüm bölümleri getir ve görüntüle
    departments = Department.query.all()
    return render_template('departments.html', departments=departments)

# Dersler sayfası
@app.route('/courses', methods=['GET', 'POST'])
@admin_required  # Sadece adminler ders ekleyip silebilir
def courses():
    """
    Ders yönetim sayfası
    GET: Ders listesini göster
    POST: Yeni ders ekle
    """
    if request.method == 'POST':
        # Form verilerini al
        code = request.form.get('code')
        name = request.form.get('name')
        department_id = request.form.get('department_id')
        instructor_id = request.form.get('instructor_id') if request.form.get('instructor_id') else None
        semester = request.form.get('semester', 1)
        
        # Aynı kodla başka ders var mı kontrol et
        if Course.query.filter_by(code=code).first():
            flash('Bu ders kodu zaten kullanımda!', 'error')
            return redirect(url_for('courses'))
        
        # Yeni ders oluştur ve kaydet
        course = Course(
            code=code, 
            name=name, 
            department_id=department_id,
            instructor_id=instructor_id,
            semester=semester
        )
        db.session.add(course)
        db.session.commit()
        
        flash('Ders başarıyla eklendi!', 'success')
        return redirect(url_for('courses'))
    
    # Gerekli verileri getir ve görüntüle
    courses = Course.query.all()
    departments = Department.query.all()
    instructors = User.query.filter_by(role='instructor').all()
    return render_template('courses.html', courses=courses, departments=departments, instructors=instructors)

# Derslikler sayfası
@app.route('/classrooms', methods=['GET', 'POST'])
@admin_required  # Sadece adminler derslik ekleyip silebilir
def classrooms():
    """
    Derslik yönetim sayfası
    GET: Derslik listesini göster
    POST: Yeni derslik ekle
    """
    if request.method == 'POST':
        # Form verilerini al
        code = request.form.get('code')
        capacity = request.form.get('capacity')
        
        # Aynı kodla başka derslik var mı kontrol et
        if Classroom.query.filter_by(code=code).first():
            flash('Bu derslik kodu zaten kullanımda!', 'error')
            return redirect(url_for('classrooms'))
        
        # Yeni derslik oluştur ve kaydet
        classroom = Classroom(code=code, capacity=capacity)
        db.session.add(classroom)
        db.session.commit()
        
        flash('Derslik başarıyla eklendi!', 'success')
        return redirect(url_for('classrooms'))
    
    # Tüm derslikleri getir ve görüntüle
    classrooms = Classroom.query.all()
    return render_template('classrooms.html', classrooms=classrooms)

# Kullanıcılar sayfası
@app.route('/users', methods=['GET', 'POST'])
@admin_required  # Sadece adminler kullanıcı ekleyip silebilir
def users():
    """
    Kullanıcı yönetim sayfası
    GET: Kullanıcı listesini göster
    POST: Yeni kullanıcı ekle
    """
    if request.method == 'POST':
        # Form verilerini al
        username = request.form.get('username')
        password = request.form.get('password')
        role = request.form.get('role')
        name = request.form.get('name')
        department_id = request.form.get('department_id') if request.form.get('department_id') else None
        extra_info = request.form.get('extra_info')
        
        # Aynı kullanıcı adıyla başka kullanıcı var mı kontrol et
        if User.query.filter_by(username=username).first():
            flash('Bu kullanıcı adı zaten kullanımda!', 'error')
            return redirect(url_for('users'))
        
        # Yeni kullanıcı oluştur ve kaydet
        user = User(
            username=username, 
            password=password, 
            role=role,
            name=name,
            department_id=department_id
        )
        
        db.session.add(user)
        db.session.commit()
        
        flash('Kullanıcı başarıyla eklendi!', 'success')
        return redirect(url_for('users'))
    
    # Gerekli verileri getir ve görüntüle
    users = User.query.all()
    departments = Department.query.all()
    return render_template('users.html', users=users, departments=departments)

# Kullanıcı silme endpoint'i
@app.route('/users/delete/<int:user_id>', methods=['POST'])
@admin_required
def delete_user(user_id):
    """
    Belirtilen ID'ye sahip kullanıcıyı siler
    :param user_id: Silinecek kullanıcının ID'si
    """
    try:
        # Kendini silmeye çalışıyor mu kontrolü
        if current_user.id == user_id:
            flash('Kendi hesabınızı silemezsiniz!', 'error')
            return redirect(url_for('users'))
        
        # Kullanıcıyı bul
        user = User.query.get_or_404(user_id)
        
        # Admin silinmeye çalışılıyor ve başka admin var mı kontrolü
        if user.role == 'admin':
            admin_count = User.query.filter_by(role='admin').count()
            if admin_count <= 1:
                flash('Son admin kullanıcıyı silemezsiniz!', 'error')
                return redirect(url_for('users'))
        
        # Kullanıcıyı sil
        db.session.delete(user)
        db.session.commit()
        flash('Kullanıcı başarıyla silindi!', 'success')
    except Exception as e:
        # Hata durumunda logla ve kullanıcıya bildir
        print(f"\n=== Hata ===")
        print(f"Hata mesajı: {str(e)}")
        print("============\n")
        flash('Kullanıcı silinirken bir hata oluştu!', 'error')
    
    return redirect(url_for('users'))

# Ders programı görüntüleme sayfası
@app.route('/view_schedule')
@login_required  # Sadece giriş yapmış kullanıcılar görebilir
def view_schedule():
    """
    Ders programını görüntüleme sayfası
    Tüm dersleri, derslikleri ve ders programını gösterir
    """
    # Haftanın günleri
    days = ['Pazartesi', 'Salı', 'Çarşamba', 'Perşembe', 'Cuma']
    
    # Veritabanından gerekli verileri çek
    schedule_items = Schedule.query.all()
    courses = Course.query.order_by(Course.code).all()  # Dersleri kod sırasına göre sırala
    classrooms = Classroom.query.order_by(Classroom.code).all()  # Derslikleri kod sırasına göre sırala
    
    # Debug için konsola bilgi yazdır
    print("\n=== Debug Bilgileri ===")
    print(f"Toplam ders sayısı: {len(courses)}")
    print(f"Toplam derslik sayısı: {len(classrooms)}")
    
    print("\nDersler:")
    for course in courses:
        print(f"- {course.code} - {course.name} (ID: {course.id})")
    
    print("\nDerslikler:")
    for classroom in classrooms:
        print(f"- {classroom.code} (Kapasite: {classroom.capacity}, ID: {classroom.id})")
    print("=====================\n")
    
    # Şablonu render et
    return render_template('view_schedule.html',
                         schedule_items=schedule_items,
                         courses=courses,
                         classrooms=classrooms,
                         days=days)

# Program ekle endpoint'i
@app.route('/schedule/add', methods=['GET', 'POST'])
@admin_required  # Sadece adminler program ekleyebilir
def add_schedule():
    """
    Ders programına yeni bir ders ekler
    GET: Yönlendirme yapar
    POST: Yeni programı kaydeder
    """
    try:
        if request.method == 'GET':
            return redirect(url_for('view_schedule'))
        
        # Form verilerini al
        course_id = request.form.get('course_id')
        classroom_id = request.form.get('classroom_id')
        day = request.form.get('day')
        start_time = request.form.get('start_time')
        end_time = request.form.get('end_time')

        # Debug için form verilerini yazdır
        print(f"\n=== Form Verileri ===")
        print(f"Ders ID: {course_id}")
        print(f"Derslik ID: {classroom_id}")
        print(f"Gün: {day}")
        print(f"Başlangıç: {start_time}")
        print(f"Bitiş: {end_time}")
        print("==================\n")

        # Seçilen dersin öğretim üyesini bul
        course = Course.query.get(course_id)
        if course and course.instructor_id:
            instructor = User.query.get(course.instructor_id)
            print(f"Ders öğretim üyesi: {instructor.name if instructor else 'Atanmamış'}")
            
            # Bu gün ve saatte öğretim üyesinin başka dersi var mı kontrol et
            instructor_conflicts = Schedule.query.join(Course).filter(
                Schedule.day == day,
                Schedule.start_time < end_time,
                Schedule.end_time > start_time,
                Course.instructor_id == course.instructor_id
            ).all()
            
            if instructor_conflicts:
                conflict_details = []
                for conflict in instructor_conflicts:
                    conflict_course = Course.query.get(conflict.course_id)
                    conflict_classroom = Classroom.query.get(conflict.classroom_id)
                    conflict_details.append(f"{conflict_course.code} ({conflict_classroom.code}, {conflict.start_time}-{conflict.end_time})")
                
                # Öğretim üyesi çakışması varsa uyar
                conflict_message = ", ".join(conflict_details)
                flash(f'Öğretim üyesi ({instructor.name}) bu saatte başka bir derste meşgul: {conflict_message}', 'error')
                return redirect(url_for('view_schedule'))

        # Seçilen derslik ve zamanda başka ders var mı kontrol et
        classroom_conflicts = Schedule.query.filter(
            Schedule.day == day,
            Schedule.start_time < end_time,
            Schedule.end_time > start_time,
            Schedule.classroom_id == classroom_id
        ).all()
        
        if classroom_conflicts:
            # Derslik çakışması varsa uyar
            conflict_details = []
            for conflict in classroom_conflicts:
                conflict_course = Course.query.get(conflict.course_id)
                conflict_details.append(f"{conflict_course.code} ({conflict.start_time}-{conflict.end_time})")
            
            conflict_message = ", ".join(conflict_details)
            conflict_classroom = Classroom.query.get(classroom_id)
            flash(f'Derslik {conflict_classroom.code} bu saatte dolu: {conflict_message}', 'error')
            return redirect(url_for('view_schedule'))
        
        # Yeni program öğesi oluştur ve kaydet
        schedule_item = Schedule(
            course_id=course_id,
            classroom_id=classroom_id,
            day=day,
            start_time=start_time,
            end_time=end_time
        )
        
        db.session.add(schedule_item)
        db.session.commit()
        
        flash('Ders programı başarıyla güncellendi!', 'success')
        
    except Exception as e:
        # Hata durumunda logla ve kullanıcıya bildir
        print(f"\n=== Hata ===")
        print(f"Hata mesajı: {str(e)}")
        print("============\n")
        flash('Ders programı eklenirken bir hata oluştu!', 'error')
        
    return redirect(url_for('view_schedule'))

# Program sil endpoint'i
@app.route('/schedule/delete/<int:schedule_id>', methods=['POST'])
@admin_required  # Sadece adminler program silebilir
def delete_schedule(schedule_id):
    """
    Belirtilen ID'ye sahip program öğesini siler
    :param schedule_id: Silinecek program öğesinin ID'si
    """
    try:
        # Program öğesini bul ve sil
        schedule_item = Schedule.query.get_or_404(schedule_id)
        db.session.delete(schedule_item)
        db.session.commit()
        flash('Program öğesi başarıyla silindi!', 'success')
    except Exception as e:
        # Hata durumunda logla ve kullanıcıya bildir
        flash('Program öğesi silinirken bir hata oluştu!', 'error')
        print(f"\n=== Hata ===")
        print(f"Hata mesajı: {str(e)}")
        print("============\n")
    
    return redirect(url_for('view_schedule'))

# Bölüm silme endpoint'i
@app.route('/departments/delete/<int:department_id>', methods=['POST'])
@admin_required  # Sadece adminler bölüm silebilir
def delete_department(department_id):
    """
    Belirtilen ID'ye sahip bölümü siler
    :param department_id: Silinecek bölümün ID'si
    """
    try:
        # Bölümün kullanıldığı dersleri kontrol et
        courses_in_department = Course.query.filter_by(department_id=department_id).count()
        users_in_department = User.query.filter_by(department_id=department_id).count()
        
        # İlişkili kayıtlar varsa silme
        if courses_in_department > 0 or users_in_department > 0:
            flash(f'Bu bölüm silinemez: {courses_in_department} ders ve {users_in_department} kullanıcı bu bölüme bağlı!', 'error')
            return redirect(url_for('departments'))
            
        # Bölümü bul ve sil
        department = Department.query.get_or_404(department_id)
        db.session.delete(department)
        db.session.commit()
        flash('Bölüm başarıyla silindi!', 'success')
    except Exception as e:
        # Hata durumunda logla ve kullanıcıya bildir
        flash('Bölüm silinirken bir hata oluştu!', 'error')
        print(f"\n=== Hata ===")
        print(f"Hata mesajı: {str(e)}")
        print("============\n")
    
    return redirect(url_for('departments'))

# Ders silme endpoint'i
@app.route('/courses/delete/<int:course_id>', methods=['POST'])
@admin_required  # Sadece adminler ders silebilir
def delete_course(course_id):
    """
    Belirtilen ID'ye sahip dersi siler
    :param course_id: Silinecek dersin ID'si
    """
    try:
        # Dersin kullanıldığı program öğeleri var mı kontrol et
        schedule_count = Schedule.query.filter_by(course_id=course_id).count()
        
        # İlişkili kayıtlar varsa silme
        if schedule_count > 0:
            flash(f'Bu ders silinemez: {schedule_count} program öğesi bu derse bağlı!', 'error')
            return redirect(url_for('courses'))
            
        # Dersi bul ve sil
        course = Course.query.get_or_404(course_id)
        db.session.delete(course)
        db.session.commit()
        flash('Ders başarıyla silindi!', 'success')
    except Exception as e:
        # Hata durumunda logla ve kullanıcıya bildir
        flash('Ders silinirken bir hata oluştu!', 'error')
        print(f"\n=== Hata ===")
        print(f"Hata mesajı: {str(e)}")
        print("============\n")
    
    return redirect(url_for('courses'))

# Ders düzenleme endpoint'i
@app.route('/courses/edit/<int:course_id>', methods=['GET', 'POST'])
@admin_required  # Sadece adminler ders düzenleyebilir
def edit_course(course_id):
    """
    Belirtilen ID'ye sahip dersi düzenler
    :param course_id: Düzenlenecek dersin ID'si
    GET: Düzenleme formunu göster
    POST: Değişiklikleri kaydet
    """
    # Düzenlenecek dersi getir
    course = Course.query.get_or_404(course_id)
    
    if request.method == 'POST':
        try:
            # Form verilerini al
            name = request.form.get('name')
            department_id = request.form.get('department_id')
            instructor_id = request.form.get('instructor_id') if request.form.get('instructor_id') else None
            semester = request.form.get('semester', 1)
            
            # Dersi güncelle
            course.name = name
            course.department_id = department_id
            course.instructor_id = instructor_id
            course.semester = semester
            
            db.session.commit()
            flash('Ders başarıyla güncellendi!', 'success')
            return redirect(url_for('courses'))
        except Exception as e:
            # Hata durumunda logla ve kullanıcıya bildir
            flash('Ders güncellenirken bir hata oluştu!', 'error')
            print(f"\n=== Hata ===")
            print(f"Hata mesajı: {str(e)}")
            print("============\n")
    
    # Formda kullanılacak verileri getir
    departments = Department.query.all()
    instructors = User.query.filter_by(role='instructor').all()
    return render_template('edit_course.html', course=course, departments=departments, instructors=instructors)

# Derslik silme endpoint'i
@app.route('/classrooms/delete/<int:classroom_id>', methods=['POST'])
@admin_required  # Sadece adminler derslik silebilir
def delete_classroom(classroom_id):
    """
    Belirtilen ID'ye sahip dersliği siler
    :param classroom_id: Silinecek dersliğin ID'si
    """
    try:
        # Dersliğin kullanıldığı program öğeleri var mı kontrol et
        schedule_count = Schedule.query.filter_by(classroom_id=classroom_id).count()
        
        # İlişkili kayıtlar varsa silme
        if schedule_count > 0:
            flash(f'Bu derslik silinemez: {schedule_count} program öğesi bu dersliğe bağlı!', 'error')
            return redirect(url_for('classrooms'))
            
        # Dersliği bul ve sil
        classroom = Classroom.query.get_or_404(classroom_id)
        db.session.delete(classroom)
        db.session.commit()
        flash('Derslik başarıyla silindi!', 'success')
    except Exception as e:
        # Hata durumunda logla ve kullanıcıya bildir
        flash('Derslik silinirken bir hata oluştu!', 'error')
        print(f"\n=== Hata ===")
        print(f"Hata mesajı: {str(e)}")
        print("============\n")
    
    return redirect(url_for('classrooms'))

# Derslik düzenleme endpoint'i
@app.route('/classrooms/edit/<int:classroom_id>', methods=['GET', 'POST'])
@admin_required  # Sadece adminler derslik düzenleyebilir
def edit_classroom(classroom_id):
    """
    Belirtilen ID'ye sahip dersliği düzenler
    :param classroom_id: Düzenlenecek dersliğin ID'si
    GET: Düzenleme formunu göster
    POST: Değişiklikleri kaydet
    """
    # Düzenlenecek dersliği getir
    classroom = Classroom.query.get_or_404(classroom_id)
    
    if request.method == 'POST':
        try:
            # Form verilerini al
            capacity = request.form.get('capacity')
            
            # Dersliği güncelle
            classroom.capacity = capacity
            
            db.session.commit()
            flash('Derslik başarıyla güncellendi!', 'success')
            return redirect(url_for('classrooms'))
        except Exception as e:
            # Hata durumunda logla ve kullanıcıya bildir
            flash('Derslik güncellenirken bir hata oluştu!', 'error')
            print(f"\n=== Hata ===")
            print(f"Hata mesajı: {str(e)}")
            print("============\n")
    
    return render_template('edit_classroom.html', classroom=classroom)

# Ders programını Excel'e aktarma endpoint'i
@app.route('/export_schedule', methods=['GET'])
@admin_required  # Sadece adminler programı dışa aktarabilir
def export_schedule():
    """
    Mevcut ders programını Excel formatında dışa aktarır
    """
    try:
        # Excel çalışma kitabı oluştur
        wb = Workbook()
        ws = wb.active
        ws.title = "Ders Programı"
        
        # Sütun genişliklerini ayarla
        ws.column_dimensions['A'].width = 15  # Günler için
        for col in range(2, 6):  # 1-4 sınıflar için
            ws.column_dimensions[chr(64 + col)].width = 30
            
        # Haftanın günleri
        days = ['Pazartesi', 'Salı', 'Çarşamba', 'Perşembe', 'Cuma']
        
        # Stil tanımları
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        day_font = Font(bold=True)
        day_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        day_alignment = Alignment(horizontal='center', vertical='center')
        
        cell_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # İnce kenarlık stili
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Başlık satırını hazırla - Sınıf seviyelerini ekle
        ws.cell(row=1, column=1, value="Gün/Sınıf").font = header_font
        ws.cell(row=1, column=1).fill = header_fill
        ws.cell(row=1, column=1).alignment = header_alignment
        ws.cell(row=1, column=1).border = thin_border
        
        # Sınıf seviyelerini başlıklara ekle (1. Sınıf, 2. Sınıf, vb.)
        for grade in range(1, 5):  # 1-4. sınıflar
            cell = ws.cell(row=1, column=grade+1, value=f"{grade}. Sınıf")
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
            
        # Gün satırlarını ekle
        for row, day in enumerate(days, start=2):
            # Gün adını ekle
            cell = ws.cell(row=row, column=1, value=day)
            cell.font = day_font
            cell.fill = day_fill
            cell.alignment = day_alignment
            cell.border = thin_border
            
            # Her sınıf seviyesi için bu günde olan programları bul
            for grade in range(1, 5):  # 1-4. sınıflar
                cell = ws.cell(row=row, column=grade+1, value="")
                cell.border = thin_border
                cell.alignment = cell_alignment
                
                # Bu sınıfın yarıyıllarını belirle (her sınıf 2 yarıyıl içerir)
                # 1. sınıf: 1-2, 2. sınıf: 3-4, 3. sınıf: 5-6, 4. sınıf: 7-8
                first_semester = (grade - 1) * 2 + 1
                second_semester = first_semester + 1
                semesters = [first_semester, second_semester]
                
                # Bu gün ve sınıf seviyesinde olan dersleri bul
                # BLM ve YZM bölümlerini birlikte göster
                blm_dept = Department.query.filter_by(code='BLM').first()
                yzm_dept = Department.query.filter_by(code='YZM').first()
                
                if blm_dept and yzm_dept:
                    # Bu sınıfın yarıyıllarındaki dersleri bul
                    blm_courses = Course.query.filter(
                        Course.semester.in_(semesters),
                        Course.department_id == blm_dept.id
                    ).all()
                    
                    yzm_courses = Course.query.filter(
                        Course.semester.in_(semesters),
                        Course.department_id == yzm_dept.id
                    ).all()
                    
                    # Tüm kurs ID'lerini birleştir
                    course_ids = [course.id for course in blm_courses + yzm_courses]
                    
                    if course_ids:
                        # Bu günde ve bu kurslarda olan programları bul
                        schedule_items = Schedule.query.filter(
                            Schedule.day == day,
                            Schedule.course_id.in_(course_ids)
                        ).order_by(Schedule.start_time).all()
                        
                        # Program varsa hücreye ekle
                        if schedule_items:
                            cell_text = []
                            for item in schedule_items:
                                course = Course.query.get(item.course_id)
                                classroom = Classroom.query.get(item.classroom_id)
                                instructor = User.query.get(course.instructor_id) if course.instructor_id else None
                                
                                # Dersin yarıyılını da ekle
                                dept_code = Department.query.get(course.department_id).code if course.department_id else ''
                                
                                course_info = (
                                    f"{course.code} - {course.name} ({dept_code}, {course.semester}. Yarıyıl)\n"
                                    f"Derslik: {classroom.code if classroom else 'Belirtilmemiş'}\n"
                                    f"Saat: {item.start_time}-{item.end_time}"
                                )
                                
                                if instructor:
                                    course_info += f"\nÖğr. Üyesi: {instructor.name}"
                                    
                                cell_text.append(course_info)
                            
                            cell.value = "\n\n".join(cell_text)
            
            # Satır yüksekliğini ayarla
            ws.row_dimensions[row].height = 150
            
        # Geçici dosya oluştur ve Excel'i kaydet
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
            wb.save(tmp.name)
            tmp_path = tmp.name
            
        # Excel dosyasını kullanıcıya gönder
        return send_file(
            tmp_path,
            as_attachment=True,
            download_name='ders_programi.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        # Hata durumunda logla ve kullanıcıya bildir
        flash('Ders programı dışa aktarılırken bir hata oluştu!', 'error')
        print(f"\n=== Hata ===")
        print(f"Hata mesajı: {str(e)}")
        print("============\n")
        return redirect(url_for('view_schedule'))

# Uygulama başlangıç kontrollerini yap ve sunucuyu başlat
if __name__ == '__main__':
    """
    Uygulama başlatıldığında çalışır
    - Veritabanı tabloları oluşturulur (yoksa)
    - Admin kullanıcısı oluşturulur (yoksa)
    - Flask geliştirme sunucusu başlatılır
    """
    with app.app_context():
        # Veritabanı tablolarını oluştur
        db.create_all()
        
        # Eksik sütunları ekle (migrasyon)
        try:
            # Course tablosunda instructor_id sütunu var mı kontrol et
            inspector = inspect(db.engine)
            
            # Course tablosuna instructor_id ekle
            if 'instructor_id' not in [c['name'] for c in inspector.get_columns('courses')]:
                with db.engine.begin() as conn:
                    conn.execute(text("ALTER TABLE courses ADD COLUMN instructor_id INTEGER REFERENCES users(id)"))
                print("courses tablosuna instructor_id sütunu eklendi.")
            
            # Diğer eksik sütunları da kontrol et ve ekle
            if 'semester' not in [c['name'] for c in inspector.get_columns('courses')]:
                with db.engine.begin() as conn:
                    conn.execute(text("ALTER TABLE courses ADD COLUMN semester INTEGER DEFAULT 1"))
                print("courses tablosuna semester sütunu eklendi.")
        except Exception as e:
            print(f"Migrasyon hatası: {str(e)}")
        
        # Admin kullanıcısı oluştur (yoksa)
        admin = User.query.filter_by(username='admin').first()
        if not admin:
            admin = User(username='admin', password='admin123', role='admin', name='Sistem Yöneticisi')
            db.session.add(admin)
            db.session.commit()
            print("Admin kullanıcısı oluşturuldu. Kullanıcı adı: admin, Şifre: admin123")
    
    # Geliştirme sunucusunu başlat
    app.run(debug=True) 